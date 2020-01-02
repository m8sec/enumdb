from openpyxl import Workbook

class CreateXLSX():
    def __init__(self, outfile, host, dbtype):
        self.outfile = outfile
        self.create_workbook()
        self.create_overview(host, dbtype)

    def create_workbook(self):
        self.wb = Workbook()

    def save_workbook(self, filename):
        self.wb.save(filename)

    def create_overview(self, host, dbtype):
        # Create enum overview on sheet1
        self.ws0 = self.wb.active
        self.ws0.title = "Overview"
        self.ws0['A1'] = "Target(s):"
        self.ws0['B1'] = host
        self.ws0['A2'] = "DB Type:"
        self.ws0['B2'] = dbtype
        self.ws0['A4'] = "Database"
        self.ws0['B4'] = "Table"
        self.ws0['C4'] = "Server"
        self.sheet1_row = 5

    def addto_overview(self, db, table, keyword):
        # Add db,table,and keyword to sheet1
        self.ws0.cell(row=self.sheet1_row, column=1, value=str(db))
        self.ws0.cell(row=self.sheet1_row, column=2, value=str(table))
        self.ws0.cell(row=self.sheet1_row, column=3, value=str(keyword))
        self.sheet1_row += 1

    def create_sheet(self, db, table, columns, data, host):
        # Create new sheet for each table and add table data
        ws = self.wb.create_sheet(table[0:15])
        row_count = 1
        col_count = 1
        ws['A1'] = "[+] Table: {}   Database: {}   Server: {}".format(table, db, host)
        row_count += 1
        for col in columns:
            try:
                # Error handling while writing data
                ws.cell(row=row_count, column=col_count, value=str(col))
            except:
                ws.cell(row=row_count, column=col_count, value="Failed to write data")
            col_count += 1
        col_count = 1
        row_count += 1
        for row in data:
            for item in row:
                try:
                    # Error handing while writing data (Encrypted characters)
                    ws.cell(row=row_count, column=col_count, value=str(item))
                except:
                    ws.cell(row=row_count, column=col_count, value="Failed to write data")
                col_count += 1
            col_count = 1
            row_count += 1
        self.save_workbook(self.outfile)